from openpyxl import load_workbook
from datetime import datetime
from decimal import Decimal
import re

from ofxstatement.parser import StatementParser
from ofxstatement.plugin import Plugin
from ofxstatement import statement

transactions_list_section_header = '1.1 Информация о совершенных и исполненных сделках на конец отчетного периода'
next_section_after_transactions_list_header = '1.3 Сделки за расчетный период, обязательства из которых прекращены не в результате исполнения'
cashflow_section_header = '2. Операции с денежными средствами'
next_section_after_cashflow_header = '3.1 Движение по ценным бумагам инвестора'
securities_list_section_header = '4.1 Информация о ценных бумагах'
next_section_after_securities_list_header = '4.2 Информация об инструментах, не квалифицированных в качестве ценной бумаги'

transactions_list_columns = [
    "Номерсделки",
    "Номерпоручения",
    "Датазаключения",
    "Время",
    "Видсделки",
    "Сокращенноенаименованиеактива",
    "Кодактива",
    "Ценазаединицу",
    "Валютацены",
    "Количество",
    "Сумма(безНКД)",
    "НКД",
    "Суммасделки",
    "Валютарасчетов",
    "Комиссияброкера",
]

cashflow_columns = [
    "Датаисполнения",
    "Операция",
    "Суммазачисления",
    "Суммасписания",
    "Примечание"
]

securities_list_columns = [
    "Сокращенноенаименованиеактива",
    "Кодактива"
]

trntype_mapping = {
    # Buy/sell transactions
    'Покупка': 'BUYSTOCK',
    'Продажа': 'SELLSTOCK',
    # Don't import REPO deals as individual transactions, they bloat transactions
    # log, making it hard to reconcile. Any charges or premiums
    # would be taken from cashflow summary (less things to look at).
    'РЕПО 1 Покупка': '--Skip--',
    'РЕПО 2 Покупка': '--Skip--',
    'РЕПО 1 Продажа': '--Skip--',
    'РЕПО 2 Продажа': '--Skip--',

    # Cashflow transactions
    'Пополнение счета': 'DEBIT',
    'Вывод средств': 'CREDIT',
    'Комиссия по тарифу': 'SRVCHG',

    # ignore buy/sell and fees summary, they are imported as separate transactions
    'Покупка/продажа': "--Skip--",
    'Комиссия за сделки': '--Skip--',

    # not sure what would be the best way to treat these transactions
    # in theory you can get some money from it :)
    'РЕПО': 'DEBIT',

    'Выплата дивидендов': 'DIV',
    'Выплата купонов': 'DIV',

    # if dividends get payed while you're holding short position you need to compensate
    'Возмещение дохода по дивидендам - списание': 'FEE',

    'Налог': 'OTHER',
    'Налог (купонный доход)': 'OTHER',
    'Налог (дивиденды)': 'OTHER',
}

currency_exchange_securities_mapping = {
    'USDRUB_': {
        'currency_pair': ('USD', 'RUB'),
    },
    'EURRUB_': {
        'currency_pair': ('EUR', 'RUB'),
    },
}

ticker_transform_rules = [
    {
        # EU stocks are listed as ticker@country in Tinkoff Broker
        # Yahoo Finance however expects ticker.country format
        'ticker_pattern': '@',
        'currency': 'EUR',
        'transform_to': '.',
    },
    {
        # Certain MOEX-based ETF's, namely SBSP, TMOS (could be more, haven't tested all of them)
        # are not working well in yahoo finance - they don't follow <ticker>.ME pattern and the data
        # is not accurate. Instead those could be taken from MOEX API with original ticker (replace TMOS with your ticker):
        # RUB: https://iss.moex.com/iss/engines/stock/markets/shares/boards/TQTF/securities/TMOS.xml?iss.meta=off&iss.only=securities
        # USD: https://iss.moex.com/iss/engines/stock/markets/shares/boards/TQTD/securities/TMOS.xml?iss.meta=off&iss.only=securities
        # take PREVPRICE and PREVDATE from response
        #
        # Additionally, if ticker name is too long, it's most likely ticker of bond that also needs a very different
        # price lookup API (haven't figured it out yet though).
        'ticker_pattern': r'(SBSP|TMOS|\w{6,})',
        'currency': 'RUB',
        'transform_to': r'\1',
    },
    {
        # RUB stocks can be looked up on Yahoo Finance when .ME is appended to ticker
        'ticker_pattern': r'(\w+)',
        'currency': 'RUB',
        'transform_to': r'\1.ME',
    },
    {
        # Certain securities (ex: TCS Group) could be purchased on MOEX in RUB, but divs they
        # pay are in USD. Adding .ME to such stocks as if they were RUB stocks.
        'ticker_pattern': r'(TCSG)',
        'currency': 'USD',
        'transform_to': r'\1.ME',
    },
    {
        # Certain US stocks have dot (.) in the name, ex: BRK.B. Yahoo Finance
        # expects it to be dash (-).
        'ticker_pattern': r'\.',
        'currency': 'USD',
        'transform_to': '-',
    },
]

class TinkoffBrokerExcelStatementParser(StatementParser):
    statement = None

    def __init__(self, fin: str, currency: str):
        super().__init__()
        self.statement = statement.Statement()
        self.statement.currency = currency
        
        self.workbook = load_workbook(fin, read_only=True)
        self.worksheet = self.workbook.worksheets[0]

        self.iterator = self.worksheet.iter_rows()
        self.row_number = 0
        self.security_id_by_full_name = self.get_security_id_by_full_name()
        self.iterator.close()

        self.iterator = self.worksheet.iter_rows()
        self.row_number = 0

    def split_records(self):
        try:
            self.skip_until_section_row(transactions_list_section_header)

            self.transactions_list_columns_mapping = dict.fromkeys(transactions_list_columns, -1)
            self.populate_columns_mapping(self.next_row(), self.transactions_list_columns_mapping)

            print_rows = False

            while (row := self.next_row()) is not None:
                first_col_value = self.remove_whitespace(row[0])
                if first_col_value == self.remove_whitespace("Номер сделки"):
                    continue # header row (in between pages)

                if first_col_value == self.remove_whitespace("1.2 Информация о неисполненных сделках на конец отчетного периода"):
                    # executed transactions that are in progress of being fulfilled
                    # the columns are the same, but they have different offsets
                    self.transactions_list_columns_mapping = dict.fromkeys(transactions_list_columns, -1)
                    self.populate_columns_mapping(self.next_row(), self.transactions_list_columns_mapping)
                    continue

                if first_col_value == self.remove_whitespace(next_section_after_transactions_list_header):
                    print(f"INFO: Found next section at row {self.row_number}")
                    break
                
                mapped_row = self.map_row_to_columns(row, self.transactions_list_columns_mapping)
                yield ('buy-sells', mapped_row)

            print("INFO: Finished parsing buy-sells transactions, now extracting dividends")

            self.skip_until_section_row(cashflow_section_header)
            self.next_row() # skip header row
            
            cashflow_summary_currencies = []
            while True:
                row = self.next_row()
                currency = row[0]
                # unfortunately the summary table terminates without an empty line
                # and directly with the cashflow table of the first currency
                # need to keep it in mind if the currency we are looking for is actually the first one
                try:
                    _ = cashflow_summary_currencies.index(currency)
                    break
                except ValueError:
                    cashflow_summary_currencies.append(currency)
                    continue
            
            try:
                requested_currency_position = cashflow_summary_currencies.index(self.statement.currency)
            except ValueError:
                print(f"WARN: Requested currency {self.statement.currency} is not present in the report")
                return

            if requested_currency_position == len(cashflow_summary_currencies) - 1:
                # for the last currency we'd loop until next section
                cashflow_section_ends_with_header = next_section_after_cashflow_header
            else:
                # otherwise we loop until we hit next currency
                cashflow_section_ends_with_header = cashflow_summary_currencies[requested_currency_position + 1]
            
            # as mentioned above, we've already stepped into the first currency cashflow section
            # but if our currency is not the first one we need to seek
            if requested_currency_position > 0:
                self.skip_until_section_row(self.statement.currency)
            
            self.cashflow_columns_mapping = dict.fromkeys(cashflow_columns, -1)
            self.populate_columns_mapping(self.next_row(), self.cashflow_columns_mapping)

            while (row := self.next_row()) is not None:
                if row[0] == cashflow_section_ends_with_header:
                    print(f"INFO: Found next section {cashflow_section_ends_with_header} at row {self.row_number}")
                    break

                mapped_row = self.map_row_to_columns(row, self.cashflow_columns_mapping)
                yield ('cashflow', mapped_row)
            
            self.iterator.close()
            self.workbook.close()
        except:
            print(f"ERR: Error processing row {self.row_number}")
            raise

    def parse_record(self, line):
        if line[0] == 'buy-sells':
            return self.parse_buy_sell_line(line[1])

        if line[0] == 'cashflow':
            return self.parse_cashflow_line(line[1])

    def parse_buy_sell_line(self, line):
        if (transactions := self.try_parse_as_currency_exchange(line)):
            return transactions

        if line['Валютарасчетов'] != self.statement.currency:
            return None

        transaction = statement.InvestStatementLine()
        transaction.id = str(line['Номерсделки'])
        transaction.date = datetime.strptime(f"{line['Датазаключения']} {line['Время']}", '%d.%m.%Y %H:%M:%S')
        transaction.security_id = self.transform_ticker(line['Кодактива'])

        transaction.unit_price = Decimal(str(line['Ценазаединицу']).replace(',', '.'))
        transaction.units = Decimal(str(line['Количество']).replace(',', '.'))
        transaction.fees = Decimal(str(line['Комиссияброкера']).replace(',', '.'))
        
        transaction.amount = Decimal(str(line['Суммасделки']).replace(',', '.'))

        transaction.memo = f"{line['Видсделки']} {line['Количество']} {line['Сокращенноенаименованиеактива']} ({line['Кодактива']}) по {line['Валютарасчетов']} {line['Ценазаединицу']}. Сумма: {line['Валютарасчетов']} {line['Суммасделки']}, комиссия {line['Валютарасчетов']} {transaction.fees}, номер сделки: {line['Номерсделки']}, номер поручения: {line['Номерпоручения']}"

        transaction.trntype = trntype_mapping.get(line['Видсделки'], None)
        if transaction.trntype is None:
            print(f"WARN: {transaction} skipped, transaction type {line['Видсделки']} can not be mapped to known type")
            return None
        
        if transaction.trntype == "--Skip--":
            return None

        if transaction.trntype.startswith("BUY"):
            transaction.trntype_detailed = "BUY"
            # buy transactions deduct money and add securities
            transaction.amount = -abs(transaction.amount) - transaction.fees
            transaction.units = abs(transaction.units)
        else:
            transaction.trntype_detailed = "SELL"
            # sell transactions do the opposite
            transaction.amount = abs(transaction.amount) - transaction.fees
            transaction.units = -abs(transaction.units)

        return transaction

    def try_parse_as_currency_exchange(self, line):
        pair_info = None
        for pair_pattern in currency_exchange_securities_mapping:
            security_full_name = line['Сокращенноенаименованиеактива']
            if re.search(pair_pattern, security_full_name):
                pair_info = currency_exchange_securities_mapping[pair_pattern]

        if pair_info is None:
            return None

        deal_currency = line['Валютарасчетов']
        if deal_currency != self.statement.currency:
            return None
        
        other_currency = list(filter(lambda x: x != deal_currency, pair_info['currency_pair']))
        if len(other_currency) > 1:
            print(f"WARN: {security_full_name} can not be mapped correctly with matching pair {pair_pattern}, can't find other currency")
            return None

        other_currency = other_currency[0]
        trntype = trntype_mapping.get(line['Видсделки'], None)

        transaction = statement.StatementLine()
        transaction.id = str(line['Номерсделки'])
        transaction.trntype = "XFER"
        transaction.bank_account_to = statement.BankAccount(self.statement.bank_id, other_currency)
        transaction.date = datetime.strptime(f"{line['Датазаключения']} {line['Время']}", '%d.%m.%Y %H:%M:%S')
        transaction.memo = f"{line['Видсделки']} {line['Количество']} {other_currency} по {line['Валютарасчетов']} {line['Ценазаединицу']}. Сумма: {line['Валютарасчетов']} {line['Суммасделки']}, номер сделки: {line['Номерсделки']}, номер поручения: {line['Номерпоручения']}"
        transaction.amount = abs(Decimal(str(line['Суммасделки']).replace(',', '.')))
        
        if trntype.startswith("BUY"):
            transaction.amount = -transaction.amount

        fees = abs(Decimal(str(line['Комиссияброкера']).replace(',', '.')))
        if fees == 0:
            return [transaction]

        fee_transaction = statement.StatementLine()
        fee_transaction.id = str(line['Номерсделки']) + '-fees'
        fee_transaction.trntype = "FEE"
        fee_transaction.date = datetime.strptime(f"{line['Датазаключения']} {line['Время']}", '%d.%m.%Y %H:%M:%S')
        fee_transaction.memo = f"Комиссия за {line['Видсделки']} {line['Количество']} {other_currency} по {line['Валютарасчетов']} {line['Ценазаединицу']}. Сумма: {line['Валютарасчетов']} {line['Суммасделки']}, номер сделки: {line['Номерсделки']}, номер поручения: {line['Номерпоручения']}"
        fee_transaction.amount = -fees
        return [transaction, fee_transaction]

    def parse_cashflow_line(self, line):
        trntype = trntype_mapping.get(line['Операция'], None)

        if trntype == "--Skip--":
            return None
        elif trntype is None:
            print(f"WARN: {line} skipped, transaction type {line['Операция']} can not be mapped to known type")
            return None
        elif trntype == "DIV":
            # return dividends as investments transaction so it gets linked to security
            transaction = statement.InvestStatementLine()
            # Dividends have security mentioned in memo column in the following format using full name:
            # Walmart-ао/ 3 шт.
            # sometimes can start with 'План;' meaning it's not fulfilled yet, not including those
            if line['Примечание'].startswith('План'):
                return None

            if not (matched_vals := re.findall("(.+?)/ \d+", line['Примечание'])):
                print(f"WARN: Unable to match dividends' full security name from {line['Примечание']}")
                return None
            
            security_full_name = matched_vals[0]
            if (security_id := self.security_id_by_full_name.get(security_full_name, None)) is None:
                print(f"WARN: Unknown security ticker {security_full_name}")
                return None

            transaction.security_id = security_id
            transaction.trntype = "INCOME"
            transaction.trntype_detailed = "DIV"
        else:
            transaction = statement.StatementLine()
            transaction.trntype = trntype
        
        transaction.date = datetime.strptime(line['Датаисполнения'], '%d.%m.%Y')
        debit = Decimal(str(line['Суммазачисления']).replace(',', '.'))
        credit = Decimal(str(line['Суммасписания']).replace(',', '.'))
        transaction.amount = debit - credit

        transaction.memo = f"{line['Операция']} {line['Примечание']}, зачислено {line['Суммазачисления']}, списано {line['Суммасписания']}, дата исполнения: {line['Датаисполнения']}"

        # report doesn't contain transaction IDs for cashflow
        transaction.id = statement.generate_transaction_id(transaction)
        return transaction
    
    def get_security_id_by_full_name(self):
        mapping = {}

        self.skip_until_section_row(securities_list_section_header)

        self.securities_list_columns_mapping = dict.fromkeys(securities_list_columns, -1)
        self.populate_columns_mapping(self.next_row(), self.securities_list_columns_mapping)

        while (row := self.next_row()) is not None:
            if row[0] == next_section_after_securities_list_header:
                print(f"INFO: Found next section at row {self.row_number}")
                break

            if row[0] == "Сокращенное наименование актива":
                continue # header row (in between pages)

            row = self.map_row_to_columns(row, self.securities_list_columns_mapping)
            security_full_name = row["Сокращенноенаименованиеактива"]
            security_id = self.transform_ticker(row["Кодактива"])
            mapping[security_full_name] = security_id

        return mapping

    def transform_ticker(self, ticker):
        for rule in ticker_transform_rules:
            if (currency_matcher := rule.get('currency', None)) is not None:
                if not re.match(currency_matcher, self.statement.currency):
                    continue
            
            if not re.findall(rule['ticker_pattern'], ticker):
                continue
            else:
                ticker = re.sub(rule['ticker_pattern'], rule['transform_to'], ticker)
                break
        
        return ticker

    def next_row(self):
        row = next(self.iterator, None)
        self.row_number += 1
        if row is None:
            return row
        
        row_mapped = tuple(map(lambda x: x.value, row))

        # sometimes there can be a 'page break' with empty row only having
        # pages counter like '23 из 28' in the right-most cells
        # check first few cells and skip to next row if all of them are empty
        for i in range(0, 50):
            if not row_mapped[i] is None:
                return row_mapped

        return self.next_row()

    def skip_until_section_row(self, section_header_value):
        while (row := self.next_row()) is not None:
            if len(row) > 0 and row[0] == section_header_value:
                print(f"INFO: Found section {section_header_value} header at row {self.row_number}")
                return
        
        print(f"ERR: Section {section_header_value} not found")
        raise f"Section {section_header_value} not found"

    def populate_columns_mapping(self, header_row, columns_mapping):
        for col_num in range(0, len(header_row)):
            value = header_row[col_num]
            if value is None:
                continue

            value = self.remove_whitespace(value) # remove all newlines and spaces

            if columns_mapping.get(value) == -1:
                # print(f"INFO: Found column {value} at {col_num}")
                columns_mapping[value] = col_num
        
        for key in columns_mapping:
            if columns_mapping[key] == -1:
                print(f"WARN: Couldn't find location of column {key}")

    def map_row_to_columns(self, row, columns_mapping):
        mapped_row = {}
        for col_name in columns_mapping:
            col_number = columns_mapping[col_name]
            mapped_row[col_name] = row[col_number]

        return mapped_row

    def remove_whitespace(self, value):
        if not isinstance(value, str):
            value = str(value)

        return re.sub('\s+', '', value)

class TinkoffBrokerExcelPlugin(Plugin):
    """Tinkoff
    """

    def get_parser(self, fin):
        if not fin.endswith('.xlsx'):
            raise f'Invalid report file {fin}. Expected .xslx'

        parser = TinkoffBrokerExcelStatementParser(fin, self.settings.get('currency'))
        parser.statement.account_id = self.settings['account']
        parser.statement.broker_id = self.settings.get('broker', 'Tinkoff Investments')
        return parser
